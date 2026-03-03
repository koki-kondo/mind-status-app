"""
Views for Mind Status API.
"""

from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from rest_framework import viewsets, permissions, status as http_status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Organization, User, StatusLog, InviteToken
from .serializers import OrganizationSerializer, UserSerializer, StatusLogSerializer
import logging

logger = logging.getLogger(__name__)

class OrganizationViewSet(viewsets.ModelViewSet):
    """組織API"""
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    permission_classes = [permissions.IsAuthenticated]


class UserViewSet(viewsets.ModelViewSet):
    """ユーザーAPI"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_permissions(self):
        """アクションごとに権限を設定"""
        public_actions = [
            'admin_register',           # 管理者登録
            'verify_invite',            # 招待トークン検証
            'set_password_with_invite', # 招待パスワード設定
            'request_password_reset',   # パスワードリセット要求
            'reset_password',           # パスワードリセット実行
        ]
        
        if self.action in public_actions:
            return [permissions.AllowAny()]
        
        return [permissions.IsAuthenticated()]
    
    def get_queryset(self):
        """組織でフィルタリング"""
        user = self.request.user
        if user.role == 'ADMIN':
            return User.objects.filter(organization=user.organization)
        return User.objects.filter(id=user.id)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        """現在ログイン中のユーザー情報を取得"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def admin_register(self, request):
        """
        管理者登録エンドポイント
        
        POST /api/users/admin_register/
        
        Request Body:
        {
            "email": "admin@example.com",
            "password": "Admin123",
            "full_name": "管理者名",
            "organization_name": "組織名",
            "org_type": "COMPANY" or "SCHOOL"
        }
        """
        from .serializers import AdminRegistrationSerializer
        
        serializer = AdminRegistrationSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # create() メソッドで組織とユーザーを作成
            user = serializer.save()
            
            return Response({
                'success': True,
                'message': '管理者登録が完了しました',
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'full_name': user.full_name,
                    'organization': user.organization.name,
                    'organization_type': user.organization.org_type
                }
            }, status=http_status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'登録中にエラーが発生しました: {str(e)}'},
                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def bulk_upload(self, request):
        """CSV一括登録（セキュリティ強化版）"""
        
        def format_serializer_errors(errors):
            """
            Serializer のエラーを読みやすい文字列に整形
            
            例:
            {'email': [ErrorDetail(string='このメールアドレスを持ったユーザーが既に存在します。', code='unique')]}
            → 'このメールアドレスを持ったユーザーが既に存在します。'
            
            複数エラー時:
            {'email': [...], 'grade': [...]}
            → 'このメールアドレスを持ったユーザーが既に存在します。 / 学年は1〜12の範囲で入力してください'
            """
            if isinstance(errors, dict):
                messages = []
                for field, error_list in errors.items():
                    if isinstance(error_list, list):
                        for error in error_list:
                            # ErrorDetail オブジェクトから文字列を抽出
                            error_str = str(error) if hasattr(error, '__str__') else error
                            # フィールド名は含めない
                            messages.append(error_str)
                    else:
                        # フィールド名は含めない
                        messages.append(str(error_list))
                return ' / '.join(messages)
            else:
                return str(errors)
        
        # 権限チェック
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'},
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        upload_file = request.FILES.get('file')
        if not upload_file:
            return Response(
                {'error': 'ファイルが選択されていません'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        # ファイル検証（validators.py）
        from .validators import validate_bulk_upload_file, validate_bulk_upload_row
        from .serializers import BulkUploadUserSerializer
        from django.db import transaction
        
        try:
            validate_bulk_upload_file(upload_file)
        except serializers.ValidationError as e:
            return Response(
                {'error': str(e)},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        # ファイル拡張子
        file_ext = upload_file.name.lower().split('.')[-1]
        
        # ファイル読み込み
        try:
            rows = []
            
            # Excel形式の場合
            if file_ext in ['xlsx', 'xls']:
                from openpyxl import load_workbook
                
                wb = load_workbook(upload_file, data_only=True)
                org_type = request.user.organization.org_type
                
                # シート選択
                if org_type == 'SCHOOL':
                    ws = wb['学校向けテンプレート'] if '学校向けテンプレート' in wb.sheetnames else wb.worksheets[0]
                else:
                    ws = wb['企業向けテンプレート'] if '企業向けテンプレート' in wb.sheetnames else (wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0])
                
                # ヘッダー取得（2行目）
                headers = [cell.value for cell in ws[2]]
                
                # データ行読み込み（3行目以降）
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if any(row):
                        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                        rows.append(row_dict)
            
            # CSV形式の場合
            else:
                import csv
                import io
                decoded_file = upload_file.read().decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))
                rows = list(csv_reader)
            
        except Exception as e:
            return Response(
                {'error': f'ファイルの読み込みに失敗しました: {str(e)}'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        # 一括登録処理（行単位でエラーハンドリング）
        success_list = []
        error_list = []
        
        for row_num, row in enumerate(rows, start=3 if file_ext in ['xlsx', 'xls'] else 2):
            try:
                # 1. ホワイトリスト検証
                validated_row = validate_bulk_upload_row(
                    row,
                    request.user.organization.org_type,
                    row_num
                )
                
                email = validated_row.get('email', '').lower().strip()
                
                # 既存ユーザーを検索（未アクティブのみ更新対象）
                existing_user = User.objects.filter(
                    email=email,
                    organization=request.user.organization,
                    is_activated=False  # 未アクティブのみ
                ).first()
                
                if existing_user:
                    # 未アクティブユーザーは更新
                    serializer = BulkUploadUserSerializer(
                        existing_user,
                        data=validated_row,
                        context={'organization': request.user.organization},  # 重複チェック用
                        partial=True
                    )
                    if not serializer.is_valid():
                        error_list.append({
                            'row': row_num,
                            'email': email,
                            'error': format_serializer_errors(serializer.errors)
                        })
                        continue
                    
                    user = serializer.save(
                        organization=request.user.organization,
                        role='USER',
                        is_activated=False,
                        is_staff=False,
                        is_superuser=False
                    )
                    
                    # 既存の招待トークンを無効化
                    InviteToken.objects.filter(user=user).update(is_used=True)
                else:
                    # 新規作成
                    serializer = BulkUploadUserSerializer(
                        data=validated_row,
                        context={'organization': request.user.organization}  # 重複チェック用
                    )
                    if not serializer.is_valid():
                        error_list.append({
                            'row': row_num,
                            'email': email,
                            'error': format_serializer_errors(serializer.errors)
                        })
                        continue
                    
                    user = serializer.save(
                        organization=request.user.organization,
                        role='USER',
                        is_activated=False,
                        is_staff=False,
                        is_superuser=False,
                        password=User.objects.make_random_password(length=12)
                    )
                
                # 4. 招待トークン生成
                invite_token = InviteToken.objects.create(
                    user=user,
                    token_type='INVITE',
                    expires_at=timezone.now() + timedelta(days=7)
                )
                
                # 5. 招待メール送信
                try:
                    from .utils.email import send_invite_email
                    from django.conf import settings
                
                    invite_url = f"{settings.FRONTEND_URL}/invite/{invite_token.token}"

                    success = send_invite_email(
                        user_email=user.email,
                        user_name=user.full_name,
                        invite_url=invite_url
                    )

                    if not success:
                        logger.warning(f'招待メール送信失敗: {user.email}')
                    
                except Exception as e:
                    logger.error(
                    f'招待メール送信中に予期しないエラー: {user.email} '
                        f'- {type(e).__name__}: {str(e)}'
                    )
                    
                success_list.append({
                    'row': row_num,
                    'email': user.email
                })
                
            except serializers.ValidationError as e:
                # バリデーションエラー
                error_detail = str(e.detail) if hasattr(e, 'detail') else str(e)
                error_list.append({
                    'row': row_num,
                    'email': row.get('email', ''),
                    'error': error_detail
                })
            
            except Exception as e:
                # 予期しないエラー
                error_list.append({
                    'row': row_num,
                    'email': row.get('email', ''),
                    'error': f'エラー: {str(e)}'
                })
        
        # 結果を返す
        return Response({
            'success_count': len(success_list),
            'error_count': len(error_list),
            'errors': error_list
        })
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.AllowAny])
    def verify_invite(self, request):
        """招待トークンを検証"""
        token = request.query_params.get('token')
        
        if not token:
            return Response(
                {'error': 'トークンが必要です'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # is_used=False を削除（is_valid()内で判定）
            invite_token = InviteToken.objects.get(
                token=token,
                token_type='INVITE'
            )
            
            # is_valid() メソッドで判定
            if not invite_token.is_valid():
                return Response(
                    {'error': 'トークンが無効または期限切れです'},
                    status=http_status.HTTP_400_BAD_REQUEST
                )
            
            # トークンに紐づくユーザー情報を返す
            user = invite_token.user
            return Response({
                'user': {
                    'email': user.email,
                    'full_name': user.full_name
                }
            })
            
        except InviteToken.DoesNotExist:
            return Response(
                {'error': '無効なトークンです'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def set_password_with_invite(self, request):
        """招待トークンを使ってパスワードを設定する"""
        token = request.data.get('token')
        password = request.data.get('password')
        
        if not token or not password:
            return Response(
                {'error': 'トークンとパスワードが必要です'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # is_used=False を削除（is_valid()内で判定）
            invite_token = InviteToken.objects.get(
                token=token,
                token_type='INVITE'
            )
            
            # is_valid() メソッドで判定
            if not invite_token.is_valid():
                return Response(
                    {'error': 'トークンが無効または期限切れです'},
                    status=http_status.HTTP_400_BAD_REQUEST
                )
            
            # パスワード強度チェック
            error = self._validate_password_strength(password)
            if error:
                return Response({'error': error}, status=http_status.HTTP_400_BAD_REQUEST)
            
            # パスワード設定
            user = invite_token.user
            user.set_password(password)
            user.is_activated = True
            user.save()
            
            # トークンを使用済みに
            invite_token.is_used = True
            invite_token.save()
            
            return Response({
                'success': True,
                'message': 'パスワードが設定されました。ログインしてください。'
            })
            
        except InviteToken.DoesNotExist:
            return Response(
                {'error': '無効なトークンです'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
    
    # ─── パスワード変更（ログイン済みユーザー） ─────────────────
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def change_password(self, request):
        """ログイン済みユーザーがパスワードを変更する"""
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        
        if not current_password or not new_password:
            return Response(
                {'error': '現在のパスワードと新しいパスワードが必要です'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        user = request.user
        
        # 現在のパスワード検証
        if not user.check_password(current_password):
            return Response(
                {'error': '現在のパスワードが正しくありません'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        # 新しいパスワード強度チェック
        error = self._validate_password_strength(new_password)
        if error:
            return Response({'error': error}, status=http_status.HTTP_400_BAD_REQUEST)
        
        user.set_password(new_password)
        user.save()
        
        return Response({'success': True, 'message': 'パスワードが変更されました'})
    
    # ─── パスワードリセット要求（メール送信） ────────────────────
    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def request_password_reset(self, request):
        """パスワード忘れ時にリセットURLをメールで送信"""
        email = request.data.get('email')
        
        if not email:
            return Response(
                {'error': 'メールアドレスが必要です'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        # ユーザーが存在するか確認（存在しなくても同じレスポンスを返す：列挙攻撃対策）
        try:
            user = User.objects.get(email=email, is_activated=True)
        except User.DoesNotExist:
            # あえて成功レスポンスを返す
            return Response({
                'success': True,
                'message': 'メールアドレスが登録されている場合、リセットリンクが送られます'
            })
        
        # リセットトークン生成
        from .models import InviteToken
        reset_token = InviteToken.objects.create(
            user=user,
            token_type='RESET',
            expires_at=timezone.now() + timedelta(hours=1)  # 1時間有効
        )
        
        # リセットメール送信（外部API障害でサービス全体が落ちないように保護）
        try:
            from .utils.email import send_password_reset_email
            from django.conf import settings

            reset_url = f"{settings.FRONTEND_URL}/reset-password/{reset_token.token}"

            success = send_password_reset_email(
                user_email=user.email,
                user_name=user.full_name,
                reset_url=reset_url
            )

            if not success:
                logger.warning(f'パスワードリセットメール送信失敗: {user.email}')
                
        except Exception as e:
            # メール送信失敗してもトークンは作成済みなので処理は続行
            logger.error(
                f'パスワードリセットメール送信中に予期しないエラー: {user.email} '
                f'- {type(e).__name__}: {str(e)}'
            )
        
        return Response({
            'success': True,
            'message': 'メールアドレスが登録されている場合、リセットリンクが送られます'
        })
    
    # ─── パスワード再設定（リセットトークン使用） ────────────────
    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def reset_password(self, request):
        """パスワードリセットトークンを使って新しいパスワードを設定する"""
        token = request.data.get('token')
        password = request.data.get('password')
        
        if not token or not password:
            return Response(
                {'error': 'トークンとパスワードが必要です'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # is_used=False を削除（is_valid()内で判定）
            reset_token = InviteToken.objects.get(
                token=token,
                token_type='RESET'
            )
            
            # is_valid() メソッドで判定
            if not reset_token.is_valid():
                return Response(
                    {'error': 'トークンが無効または期限切れです'},
                    status=http_status.HTTP_400_BAD_REQUEST
                )
            
            # パスワード強度チェック
            error = self._validate_password_strength(password)
            if error:
                return Response({'error': error}, status=http_status.HTTP_400_BAD_REQUEST)
            
            # パスワード設定
            user = reset_token.user
            user.set_password(password)
            user.save()
            
            # トークンを使用済みに
            reset_token.is_used = True
            reset_token.save()
            
            return Response({
                'success': True,
                'message': 'パスワードがリセットされました。ログインしてください。'
            })
            
        except InviteToken.DoesNotExist:
            return Response(
                {'error': '無効なトークンです'},
                status=http_status.HTTP_400_BAD_REQUEST
            )
    
    # ─── パスワード強度チェック（共通） ───────────────────────────
    def _validate_password_strength(self, password: str) -> str:
        """パスワード強度を検証。エラーメッセージを返す。問題なければ空文字列"""
        if len(password) < 8:
            return 'パスワードは8文字以上で設定してください'
        if not any(c.isupper() for c in password):
            return 'パスワードには大文字を含めてください'
        if not any(c.islower() for c in password):
            return 'パスワードには小文字を含めてください'
        if not any(c.isdigit() for c in password):
            return 'パスワードには数字を含めてください'
        return ''

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def csv_template(self, request):
        """Excelテンプレートダウンロード（学校用・企業用2シート構成）"""
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'},
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io
        
        # Excelワークブック作成
        wb = Workbook()
        
        # ─── シート1: 学校向け ───
        ws_school = wb.active
        ws_school.title = '学校向けテンプレート'
        
        # ヘッダー行（学校用）
        school_headers = ['student_number', 'full_name', 'full_name_kana', 'grade', 'class_name', 'gender', 'birth_date', 'email']
        school_headers_jp = ['学籍番号・出席番号', '氏名', 'フリガナ', '学年', '組・クラス', '性別', '生年月日', 'メールアドレス']
        
        # 日本語ヘッダー（1行目）
        for col_num, header in enumerate(school_headers_jp, 1):
            cell = ws_school.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 英語キー（2行目）
        for col_num, header in enumerate(school_headers, 1):
            cell = ws_school.cell(row=2, column=col_num, value=header)
            cell.font = Font(italic=True, color="666666")
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # サンプルデータ（学校）
        school_samples = [
            ['S2024001', '田中太郎', 'タナカタロウ', '1', 'A組', '男', '2010-04-15', 'tanaka@example.com'],
            ['S2024002', '鈴木花子', 'スズキハナコ', '1', 'A組', '女', '2010-08-22', 'suzuki@example.com'],
        ]
        for row_num, sample in enumerate(school_samples, 3):
            for col_num, value in enumerate(sample, 1):
                ws_school.cell(row=row_num, column=col_num, value=value)
        
        # 列幅調整
        ws_school.column_dimensions['A'].width = 18  # 学籍番号・出席番号
        ws_school.column_dimensions['B'].width = 15  # 氏名
        ws_school.column_dimensions['C'].width = 18  # フリガナ
        ws_school.column_dimensions['D'].width = 8   # 学年
        ws_school.column_dimensions['E'].width = 12  # 組・クラス
        ws_school.column_dimensions['F'].width = 10  # 性別
        ws_school.column_dimensions['G'].width = 15  # 生年月日
        ws_school.column_dimensions['H'].width = 25  # メールアドレス
        
        # ─── シート2: 企業向け ───
        ws_company = wb.create_sheet(title='企業向けテンプレート')
        
        # ヘッダー行（企業用）
        company_headers = ['employee_number', 'full_name', 'full_name_kana', 'department', 'position', 'gender', 'birth_date', 'email']
        company_headers_jp = ['社員番号', '氏名', 'フリガナ', '所属・部署', '役職', '性別', '生年月日', 'メールアドレス']
        
        # 日本語ヘッダー（1行目）
        for col_num, header in enumerate(company_headers_jp, 1):
            cell = ws_company.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 英語キー（2行目）
        for col_num, header in enumerate(company_headers, 1):
            cell = ws_company.cell(row=2, column=col_num, value=header)
            cell.font = Font(italic=True, color="666666")
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # サンプルデータ（企業）
        company_samples = [
            ['E001', '田中太郎', 'タナカタロウ', '営業部', '課長', '男', '1985-04-15', 'tanaka@example.com'],
            ['E002', '鈴木花子', 'スズキハナコ', '人事部', '主任', '女', '1990-08-22', 'suzuki@example.com'],
        ]
        for row_num, sample in enumerate(company_samples, 3):
            for col_num, value in enumerate(sample, 1):
                ws_company.cell(row=row_num, column=col_num, value=value)
        
        # 列幅調整
        ws_company.column_dimensions['A'].width = 15
        ws_company.column_dimensions['B'].width = 15
        ws_company.column_dimensions['C'].width = 18
        ws_company.column_dimensions['D'].width = 15
        ws_company.column_dimensions['E'].width = 12
        ws_company.column_dimensions['F'].width = 10
        ws_company.column_dimensions['G'].width = 15
        ws_company.column_dimensions['H'].width = 25
        
        # Excelファイルとして出力
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_template.xlsx"'
        
        return response
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def register_admin(self, request):
        """管理者登録API（組織とアカウントを同時作成）"""
        from .serializers import AdminRegistrationSerializer
        
        serializer = AdminRegistrationSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                user = serializer.save()
                
                return Response({
                    'success': True,
                    'message': '管理者アカウントが作成されました。ログインしてください。',
                    'user': {
                        'email': user.email,
                        'full_name': user.full_name,
                        'organization': user.organization.name,
                        'org_type': user.organization.org_type
                    }
                }, status=http_status.HTTP_201_CREATED)
                
            except Exception as e:
                return Response(
                    {'error': f'アカウント作成に失敗しました: {str(e)}'},
                    status=http_status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        return Response(
            {'errors': serializer.errors},
            status=http_status.HTTP_400_BAD_REQUEST
        )
    
    @action(detail=True, methods=['delete'], permission_classes=[permissions.IsAuthenticated])
    def delete_user(self, request, pk=None):
        """ユーザー削除API（管理者または本人のみ）"""
        try:
            user_to_delete = User.objects.get(pk=pk)
            
            # 権限チェック
            if request.user.role == 'ADMIN':
                # 管理者: 同じ組織のユーザーを削除可能（自分自身も含む）
                if user_to_delete.organization != request.user.organization:
                    return Response(
                        {'error': '他の組織のユーザーは削除できません'},
                        status=http_status.HTTP_403_FORBIDDEN
                    )
            else:
                # 一般ユーザー: 自分自身のみ削除可能
                if user_to_delete.id != request.user.id:
                    return Response(
                        {'error': '他のユーザーを削除する権限がありません'},
                        status=http_status.HTTP_403_FORBIDDEN
                    )
            
            # 削除実行
            user_email = user_to_delete.email
            user_to_delete.delete()
            
            return Response({
                'success': True,
                'message': f'ユーザー {user_email} を削除しました'
            }, status=http_status.HTTP_200_OK)
            
        except User.DoesNotExist:
            return Response(
                {'error': 'ユーザーが見つかりません'},
                status=http_status.HTTP_404_NOT_FOUND
            )


class StatusLogViewSet(viewsets.ModelViewSet):
    """ステータス記録API"""
    queryset = StatusLog.objects.all()
    serializer_class = StatusLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """ユーザーの権限に応じてフィルタリング"""
        user = self.request.user
        if user.role == 'ADMIN':
            # 管理者: 同じ組織の全ユーザーのステータス
            return StatusLog.objects.filter(user__organization=user.organization)
        # 一般ユーザー: 自分のステータスのみ
        return StatusLog.objects.filter(user=user)
    
    def perform_create(self, serializer):
        """ステータス作成時に自動的にユーザーを設定（管理者は記録不可）"""
        if self.request.user.role == 'ADMIN':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('管理者はステータスを記録できません')
        serializer.save(user=self.request.user)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def dashboard_summary(self, request):
        """管理者用ダッシュボードサマリー"""
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'}, 
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        organization = request.user.organization
        
        # 日本時間で本日の日付を取得
        import pytz
        jst = pytz.timezone('Asia/Tokyo')
        now_jst = timezone.now().astimezone(jst)
        today = now_jst.date()
        
        # デバッグログ
        logger.debug(f"Current time (JST): {now_jst}")
        logger.debug(f"Today's date: {today}")
        
        # 全ユーザー数（管理者を除外・有効化済みのみ）
        total_users = User.objects.filter(
            organization=organization,
            role='USER',
            is_activated=True  # 有効化済みユーザーのみ
        ).count()
        
        # 各ユーザーの本日の最新ステータスを取得（有効化済みのみ）
        users = User.objects.filter(
            organization=organization,
            role='USER',
            is_activated=True  # 有効化済みユーザーのみ
        )
        
        status_distribution = {
            'GREEN': 0,
            'YELLOW': 0,
            'RED': 0
        }
        today_recorded = 0
        red_alerts = 0
        
        for user in users:
            # このユーザーの本日の最新ステータス
            latest_status = StatusLog.objects.filter(
                user=user,
                created_at__date=today
            ).order_by('-created_at').first()
            
            if latest_status:
                # デバッグログ
                logger.debug(
                f"User: {user.full_name}, Status: {latest_status.status}, "
                    f"Time: {latest_status.created_at}"
                )
                
                today_recorded += 1
                status_distribution[latest_status.status] += 1
                if latest_status.status == 'RED':
                    red_alerts += 1
        
        return Response({
            'total_users': total_users,
            'today_recorded': today_recorded,
            'red_alerts': red_alerts,
            'status_distribution': status_distribution,
            'date': today.isoformat()
        })
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def alerts(self, request):
        """REDステータスのアラート一覧（最新ステータスがREDのユーザーのみ）"""
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'}, 
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        organization = request.user.organization
        
        # 日本時間で本日の日付を取得
        import pytz
        jst = pytz.timezone('Asia/Tokyo')
        now_jst = timezone.now().astimezone(jst)
        today = now_jst.date()
        
        # 一般ユーザーのみ取得（有効化済みのみ）
        users = User.objects.filter(
            organization=organization,
            role='USER',
            is_activated=True  # 有効化済みユーザーのみ
        )
        
        alerts = []
        for user in users:
            # このユーザーの本日の最新ステータス
            latest_status = StatusLog.objects.filter(
                user=user,
                created_at__date=today
            ).order_by('-created_at').first()
            
            # 最新ステータスがREDの場合のみアラートに追加
            if latest_status and latest_status.status == 'RED':
                # 所属を正しく表示
                if organization.org_type == 'COMPANY':
                    department = user.department or '-'
                else:  # SCHOOL
                    if user.grade and user.class_name:
                        department = f'{user.grade}年{user.class_name}'
                    else:
                        department = '-'
                
                alerts.append({
                    'id': str(latest_status.id),
                    'user_id': str(user.id),
                    'user_name': user.full_name,
                    'department': department,
                    'status': latest_status.status,
                    'comment': latest_status.comment,
                    'created_at': latest_status.created_at.isoformat()
                })
        
        return Response(alerts)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def user_latest_status(self, request):
        """全ユーザーの最新ステータス（管理者を除外）"""
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'}, 
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        organization = request.user.organization
        # 一般ユーザーのみ取得（有効化済みのみ）
        users = User.objects.filter(
            organization=organization,
            role='USER',
            is_activated=True  # 有効化済みユーザーのみ
        ).order_by('full_name')
        
        user_status_list = []
        for user in users:
            latest_log = StatusLog.objects.filter(user=user).order_by('-created_at').first()
            
            user_status_list.append({
                'id': str(user.id),
                'full_name': user.full_name,
                'email': user.email,
                'is_activated': user.is_activated,  # フロントエンド用に追加
                # 企業用
                'department': user.department or '',
                'position': user.position or '',
                # 学校用
                'grade': user.grade,
                'class_name': user.class_name or '',
                # ステータス
                'latest_status': latest_log.status if latest_log else None,
                'latest_comment': latest_log.comment if latest_log else None,
                'latest_date': latest_log.created_at.isoformat() if latest_log else None
            })
        
        return Response(user_status_list)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def trend_data(self, request):
        """ステータス推移データ（管理者用）- 期間指定可能"""
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'},
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        organization = request.user.organization
        
        # 期間パラメータ取得（デフォルト: 7日間）
        days = int(request.query_params.get('days', 7))
        if days not in [7, 14, 30]:
            days = 7  # 不正な値の場合は7日間
        
        # 日本時間で日付を生成
        import pytz
        jst = pytz.timezone('Asia/Tokyo')
        today_jst = timezone.now().astimezone(jst).date()
        
        trend_data = []
        
        for i in range(days - 1, -1, -1):  # N日前から今日まで
            target_date = today_jst - timedelta(days=i)
            
            # 一般ユーザーのみ取得（有効化済みのみ）
            users = User.objects.filter(
                organization=organization,
                role='USER',
                is_activated=True  # 有効化済みユーザーのみ
            )
            
            status_count = {
                'GREEN': 0,
                'YELLOW': 0,
                'RED': 0
            }
            
            # 各ユーザーのその日の最新ステータスを取得
            for user in users:
                latest_status = StatusLog.objects.filter(
                    user=user,
                    created_at__date=target_date
                ).order_by('-created_at').first()
                
                if latest_status:
                    status_count[latest_status.status] += 1
            
            trend_data.append({
                'date': target_date.isoformat(),
                'green': status_count['GREEN'],
                'yellow': status_count['YELLOW'],
                'red': status_count['RED'],
                'total': status_count['GREEN'] + status_count['YELLOW'] + status_count['RED']
            })
        
        return Response(trend_data)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def export_csv(self, request):
        """ユーザーステータスをExcel/CSV出力（管理者用・期間指定可能）"""
        if request.user.role != 'ADMIN':
            return Response(
                {'error': '管理者のみアクセス可能です'},
                status=http_status.HTTP_403_FORBIDDEN
            )
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        import pytz
        import io
        
        organization = request.user.organization
        
        # 出力形式（デフォルトはExcel）
        output_format = request.query_params.get('format', 'xlsx')
        
        # 期間パラメータ取得
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        # 日本時間のタイムゾーン
        jst = pytz.timezone('Asia/Tokyo')
        
        # 期間指定の有無で処理を分岐
        if start_date_str and end_date_str:
            # 期間指定モード: 全ステータス記録を出力
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': '日付形式が正しくありません（YYYY-MM-DD）'},
                    status=http_status.HTTP_400_BAD_REQUEST
                )
            
            # フィルタパラメータ取得
            department_filter = request.query_params.get('department')
            grade_filter = request.query_params.get('grade')
            class_filter = request.query_params.get('class')
            status_filter = request.query_params.get('status')
            search_query = request.query_params.get('search')
            
            # 指定期間内のステータス記録を取得（有効化済みユーザーのみ）
            logs = StatusLog.objects.filter(
                user__organization=organization,
                user__role='USER',
                user__is_activated=True,  # 有効化済みユーザーのみ
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            )
            
            # フィルタ適用（企業用）
            if department_filter and department_filter != 'all':
                logs = logs.filter(user__department=department_filter)
            
            # フィルタ適用（学校用）
            if grade_filter and grade_filter != 'all':
                logs = logs.filter(user__grade=int(grade_filter))
            if class_filter and class_filter != 'all':
                logs = logs.filter(user__class_name=class_filter)
            
            # 共通フィルタ
            if status_filter and status_filter != 'all':
                logs = logs.filter(status=status_filter)
            
            if search_query:
                logs = logs.filter(user__full_name__icontains=search_query)
            
            logs = logs.select_related('user').order_by('user__full_name', 'created_at')
            
            # Excel形式
            if output_format == 'xlsx':
                wb = Workbook()
                ws = wb.active
                ws.title = 'ステータス記録'
                
                # ヘッダー行（組織タイプ別）
                if organization.org_type == 'SCHOOL':
                    headers = ['氏名', '学年', '組・クラス', 'ステータス', 'コメント', '記録日時']
                else:  # COMPANY
                    headers = ['氏名', '所属・部署', '役職', 'ステータス', 'コメント', '記録日時']
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # データ行
                status_labels = {'GREEN': '健康', 'YELLOW': '注意', 'RED': '警告'}
                
                for row_num, log in enumerate(logs, 2):
                    if organization.org_type == 'SCHOOL':
                        ws.cell(row=row_num, column=1, value=log.user.full_name)
                        ws.cell(row=row_num, column=2, value=log.user.grade or '-')
                        ws.cell(row=row_num, column=3, value=log.user.class_name or '-')
                        ws.cell(row=row_num, column=4, value=status_labels.get(log.status, log.status))
                        ws.cell(row=row_num, column=5, value=log.comment or '-')
                        ws.cell(row=row_num, column=6, value=log.created_at.astimezone(jst).strftime('%Y-%m-%d %H:%M:%S'))
                    else:  # COMPANY
                        ws.cell(row=row_num, column=1, value=log.user.full_name)
                        ws.cell(row=row_num, column=2, value=log.user.department or '-')
                        ws.cell(row=row_num, column=3, value=log.user.position or '-')
                        ws.cell(row=row_num, column=4, value=status_labels.get(log.status, log.status))
                        ws.cell(row=row_num, column=5, value=log.comment or '-')
                        ws.cell(row=row_num, column=6, value=log.created_at.astimezone(jst).strftime('%Y-%m-%d %H:%M:%S'))
                
                # 列幅調整
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 20
                
                # Excelファイルとして出力
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f'user_status_{start_date}_{end_date}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response
        
        else:
            # 最新ステータスモード
            # フィルタパラメータ取得
            department_filter = request.query_params.get('department')
            grade_filter = request.query_params.get('grade')
            class_filter = request.query_params.get('class')
            status_filter = request.query_params.get('status')
            search_query = request.query_params.get('search')
            
            users = User.objects.filter(
                organization=organization,
                role='USER',
                is_activated=True  # 有効化済みユーザーのみ
            )
            
            # フィルタ適用（企業用）
            if department_filter and department_filter != 'all':
                users = users.filter(department=department_filter)
            
            # フィルタ適用（学校用）
            if grade_filter and grade_filter != 'all':
                users = users.filter(grade=int(grade_filter))
            if class_filter and class_filter != 'all':
                users = users.filter(class_name=class_filter)
            
            # 共通フィルタ
            if search_query:
                users = users.filter(full_name__icontains=search_query)
            
            users = users.order_by('full_name')
            
            # Excel形式
            if output_format == 'xlsx':
                wb = Workbook()
                ws = wb.active
                ws.title = '最新ステータス'
                
                # ヘッダー行（組織タイプ別）
                if organization.org_type == 'SCHOOL':
                    headers = ['氏名', '学年', '組・クラス', '最新ステータス', 'コメント', '記録日時']
                else:  # COMPANY
                    headers = ['氏名', '所属・部署', '役職', '最新ステータス', 'コメント', '記録日時']
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # データ行
                status_labels = {'GREEN': '健康', 'YELLOW': '注意', 'RED': '警告'}
                
                row_num = 2
                for user in users:
                    latest_log = StatusLog.objects.filter(user=user).order_by('-created_at').first()
                    
                    # ステータスフィルタを適用
                    if status_filter and status_filter != 'all':
                        if not latest_log or latest_log.status != status_filter:
                            continue
                    
                    if organization.org_type == 'SCHOOL':
                        ws.cell(row=row_num, column=1, value=user.full_name)
                        ws.cell(row=row_num, column=2, value=user.grade or '-')
                        ws.cell(row=row_num, column=3, value=user.class_name or '-')
                        ws.cell(row=row_num, column=4, value=status_labels.get(latest_log.status, '-') if latest_log else '-')
                        ws.cell(row=row_num, column=5, value=latest_log.comment if latest_log else '-')
                        ws.cell(row=row_num, column=6, value=latest_log.created_at.astimezone(jst).strftime('%Y-%m-%d %H:%M:%S') if latest_log else '-')
                    else:  # COMPANY
                        ws.cell(row=row_num, column=1, value=user.full_name)
                        ws.cell(row=row_num, column=2, value=user.department or '-')
                        ws.cell(row=row_num, column=3, value=user.position or '-')
                        ws.cell(row=row_num, column=4, value=status_labels.get(latest_log.status, '-') if latest_log else '-')
                        ws.cell(row=row_num, column=5, value=latest_log.comment if latest_log else '-')
                        ws.cell(row=row_num, column=6, value=latest_log.created_at.astimezone(jst).strftime('%Y-%m-%d %H:%M:%S') if latest_log else '-')
                    
                    row_num += 1
                
                # 列幅調整
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 20
                
                # Excelファイルとして出力
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="user_status_latest.xlsx"'
                
                return response

